"""
Defines the structure if an frtb.net configuration for a particular regulator,
reads such a configuration file and provides some common functions for retrieving
configuration items.

Copyright © 2024 frtb.net limited

Author: Alan Skea, frtb.net limited

Contact us at <info@frtb.net> or via our website at <https://frtb.net>

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU Affero General Public License as
published by the Free Software Foundation, either version 3 of the
License, or (at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU Affero General Public License for more details.

You should have received a copy of the GNU Affero General Public License
along with this program.  If not, see <https://www.gnu.org/licenses/>.
Copyright © 2024 by Alan Skea
"""

import os
import json
import numpy as np
import pandas as pd
import math
import openpyxl as xl
from typing import Dict, Any, List, Union
from pathlib import Path

import FRTBUtils as FNU


class NumpyEncoder(json.JSONEncoder):
    """JSON encoder that handles numpy and pandas types."""
    def default(self, obj):
        if isinstance(obj, (np.integer, np.int64)):
            return int(obj)
        elif isinstance(obj, (np.floating, np.float64)):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        elif isinstance(obj, pd.Series):
            return obj.tolist()
        elif isinstance(obj, pd.DataFrame):
            return obj.to_dict(orient='list')
        return super().default(obj)


class FRTBConfig(object):
    # _Regulators = [
    #     'BCBS',
    #     'UK-PRA',
    #     'EU-EBA',
    #     'SG-MAS'`
    # ]

    _configFileJSON = 'FRTBConfig_{}.json'
    _configFileExcel = 'FRTBConfig_{}.xlsx'

    # Type mapping for JSON serialization/deserialization
    _TYPE_MAP = {
        'str': str,
        'bool': bool,
        'float64': float,
        'int64': int
    }

    _FILLNA_MAP = {
        'str': '',
        'bool': False,
        'float64': 0.0,
        'int64': 0
    }

    #
    # The following describes the shape of the data for each of the keys in the config.  If the
    # config key appears in a list here then it is given the treatment appropriate for that list.
    # Possible list memberships are:
    #   'listKeys'      : This lists the keys that have a python list of values.  They may be laid out
    #                     horizontally or vertically.  Lists may have headers naming the items in either
    #                     the first row or columns depending on the layout of the list.  listKeys are
    #                     read inot pandas.Series objects.
    #   'arrayKeys'     : List of keys that are arrays
    #   'rowHdrKeys'    : List of keys that have row headers
    #   'colHdrKeys'    : List of keys that have column headers
    #   'addIndex'      : Dictionary of keys and expressions to eval to create index labels
    #   'addColumns'    : Dictionary of keys and expressions to eval to creatre column lalels
    #
    # The addIndex and addColumns keys contain special magic.  They are used to set the row and column
    # indices of a DataFrame or the index of a Series and can refer to other elements of the config data
    # for that riskClass.  The strings are evaluated in the context of the dataDict dictionary into which
    # all the config data is being loaded.  The usual use-case is to set an index from the list of buckets,
    # but in the IR risk classes, the index is set from the list of tenors.  The structure of the addIndex
    # and addColumns dictionaries is that the key is the name of the data item to which the row index or
    # column indsex is to be added, and the value is the string to be evaluated to get the index or column values.
    #
    _riskClassCongigKeyTypes = {
            'MR' : {
                'arrayKeys' : ['VegaLiquidityHorizon'],
                'colHdrKeys' : ['VegaLiquidityHorizon']
            },
            'MS_IR' : {
                'listKeys' : ['DeltaTenorRiskWeight', 'BaselCcys', 'DeltaTenors', 'VegaTenors', 'ERMIICcys'],
                'arrayKeys' : ['DeltaTenorRho'],
                'addIndex' : { 'DeltaTenorRiskWeight' : "dataDict['DeltaTenors']", 'DeltaTenorRho' : "dataDict['DeltaTenors']" },
                'addColumns' : { 'DeltaTenorRho' : "dataDict['DeltaTenors']" }
            },
            'MS_CR' : {
                'listKeys' : ['DeltaBucketRiskWeight', 'DeltaTenors', 'VegaTenors', 'CoveredBondHighQuality', 'IndexBuckets'],
                'arrayKeys' : ['Bucket', 'Gamma'],
                'colHdrKeys' : ['Bucket'],
                'addIndex' : { 'DeltaBucketRiskWeight' : "dataDict['Bucket']['Bucket'] + dataDict['Bucket']['SubBucket']",
                               'Gamma' : "dataDict['Bucket']['Bucket'].unique()" },
                'addColumns' : { 'Gamma' : "dataDict['Bucket']['Bucket'].unique()" }
            },
            'MS_CC' : {
                'listKeys' : ['DeltaBucketRiskWeight', 'DeltaTenors', 'VegaTenors'],
                'arrayKeys' : ['Bucket', 'Gamma'],
                'colHdrKeys' : ['Bucket'],
                'addIndex' : { 'DeltaBucketRiskWeight' : "dataDict['Bucket']['Bucket'] + dataDict['Bucket']['SubBucket']",
                               'Gamma' : "dataDict['Bucket']['Bucket'].unique()" },
                'addColumns' : { 'Gamma' : "dataDict['Bucket']['Bucket'].unique()" }
            },
            'MS_CS' : {
                'listKeys' : ['DeltaBucketRiskWeight', 'DeltaTenors', 'VegaTenors'],
                'arrayKeys' : ['Bucket'],
                'colHdrKeys' : ['Bucket'],
                'addIndex' : { 'DeltaBucketRiskWeight' : "dataDict['Bucket']['Bucket'] + dataDict['Bucket']['SubBucket']" }
            },
            'MS_EQ' : {
                'listKeys' : ['DeltaNameBucketRho', 'VegaTenors'],
                'arrayKeys' : ['AdvancedEconomyCountries', 'Bucket', 'DeltaBucketRiskWeight', 'Gamma'],
                'rowHdrKeys' : ['DeltaBucketRiskWeight'],
                'colHdrKeys' : ['Bucket'],
                'addIndex' : { 'DeltaNameBucketRho' : "dataDict['Bucket']['Bucket'].unique()",
                               'Gamma' : "dataDict['Bucket']['Bucket'].unique()" },
                'addColumns' : { 'DeltaBucketRiskWeight' : "dataDict['Bucket']['Bucket'] + dataDict['Bucket']['SubBucket']",
                                 'Gamma' : "dataDict['Bucket']['Bucket'].unique()" }
            },
            'MS_CM' : {
                'listKeys' : ['DeltaBucketRiskWeight', 'DeltaCommodityRho', 'DeltaTenors', 'VegaTenors'],
                'arrayKeys' : ['Bucket', 'Gamma'],
                'colHdrKeys' : ['Bucket'],
                'addIndex' : { 'DeltaBucketRiskWeight' : "dataDict['Bucket']['Bucket'] + dataDict['Bucket']['SubBucket']",
                               'DeltaCommodityRho' : "dataDict['Bucket']['Bucket'].unique()",
                               'Gamma' : "dataDict['Bucket']['Bucket'].unique()"},
                'addColumns' : { 'Gamma' : "dataDict['Bucket']['Bucket'].unique()" }
            },
            'MS_FX' : {
                'listKeys' : ['BaselCcys', 'VegaTenors', 'ERMIICcys', 'EURPegCcys'],
                'rowHdrKeys' : ['ERMIICcys']
            },
            'MD_CR' : {
                'listKeys' : ['Bucket', 'CQRiskWeight'],
                'rowHdrKeys' : ['CQRiskWeight'],
                'colHdrKeys' : ['Bucket']
            },
            'MD_CC' : {
                'listKeys' : ['CQRiskWeight'],
                'rowHdrKeys' : ['CQRiskWeight']
            },
            'MD_CS' : {
                'listKeys' : ['CQRiskWeight'],
                'arrayKeys' : ['Bucket'],
                'rowHdrKeys' : ['CQRiskWeight'],
                'colHdrKeys' : ['Bucket']
            },
            'MR_RR' : {
                'listKeys' : ['RiskWeight'],
                'arrayKeys' : ['Bucket'],
                'rowHdrKeys' : ['RiskWeight'],
                'colHdrKeys' : ['Bucket']
            },
            'CVA' : {
                'arrayKeys' : ['BA-Bucket', 'BA-RiskWeight'],
                'rowHdrKeys' : ['BA-RiskWeight'],
                'colHdrKeys' : ['BA-Bucket'],
                'addColumns' : { 'BA-RiskWeight' : "dataDict['BA-Bucket']['Bucket']" }
            },
            'CS_IR' : {
                'listKeys' : ['BaselCcys', 'DeltaTenorRiskWeight', 'DeltaTenors', 'ERMIICcys'],
                'arrayKeys' : ['DeltaTenorRho'],
                'addIndex' : { 'DeltaTenorRiskWeight' : "dataDict['DeltaTenors']", 'DeltaTenorRho' : "dataDict['DeltaTenors']" },
                'addColumns' : { 'DeltaTenorRho' : "dataDict['DeltaTenors']" }
            },
            'CS_FX' : {
                'listKeys' : ['ERMIICcys', 'EURPegCcys'],
                'rowHdrKeys' : ['ERMIICcys']
            },
            'CS_CC' : {
                'listKeys' : ['DeltaTenors', 'IndexBuckets'],
                'arrayKeys' : ['Bucket', 'Gamma', 'DeltaRiskWeight'],
                'rowHdrKeys' : ['DeltaRiskWeight'],
                'colHdrKeys' : ['Bucket'],
                'addIndex' : { 'Gamma' : "dataDict['Bucket']['Bucket'].unique()" },
                'addColumns' : { 'DeltaRiskWeight' : "dataDict['Bucket']['Bucket'] + dataDict['Bucket']['SubBucket']",
                                 'Gamma' : "dataDict['Bucket']['Bucket'].unique()" },
            },
            'CS_CR' : {
                'listKeys' : ['DeltaBucketRiskWeight'],
                'arrayKeys' : ['Bucket', 'Gamma'],
                'colHdrKeys' : ['Bucket'],
                'addIndex' :  { 'DeltaBucketRiskWeight' : "dataDict['Bucket']['Bucket'] + dataDict['Bucket']['SubBucket']",
                                 'Gamma' : "dataDict['Bucket']['Bucket'].unique()" },
                'addColumns' : { 'Gamma' : "dataDict['Bucket']['Bucket'].unique()" }
            },
            'CS_EQ' : {
                'listKeys' : ['AdvancedEconomyCountries', 'DeltaBucketRiskWeight', 'VegaBucketRiskWeight', 'DeltaBucketRho'],
                'arrayKeys' : ['Bucket', 'Gamma'],
                'colHdrKeys' : ['Bucket'],
                'addIndex' : { 'DeltaBucketRiskWeight' : "dataDict['Bucket']['Bucket'] + dataDict['Bucket']['SubBucket']",
                               'VegaBucketRiskWeight' : "dataDict['Bucket']['Bucket'] + dataDict['Bucket']['SubBucket']",
                               'DeltaBucketRho' : "dataDict['Bucket']['Bucket'].unique()",
                               'Gamma' : "dataDict['Bucket']['Bucket'].unique()",
                             },
                'addColumns' : { 'Gamma' : "dataDict['Bucket']['Bucket'].unique()" }
            },
            'CS_CM' : {
                'listKeys' : ['DeltaBucketRiskWeight'],
                'arrayKeys' : ['Bucket', 'Gamma'],
                'colHdrKeys' : ['Bucket'],
                'addIndex' : { 'DeltaBucketRiskWeight' : "dataDict['Bucket']['Bucket'] + dataDict['Bucket']['SubBucket']",
                               'Gamma' : "dataDict['Bucket']['Bucket'].unique()" },
                'addColumns' : { 'Gamma' : "dataDict['Bucket']['Bucket'].unique()" }
            }
        }

    # Everything not mentioned in these dictionaries are left as the type passed in.
    #
    _riskClassKeyDataType = {
        'MR' : {
            'VegaOptionRhoAlpha' : 'float64',
            'VegaRiskWeightSigma' : 'float64'
        },
        'MS_IR' : {
            'DeltaTenorRiskWeight' : 'float64',
            'DeltaInflationRiskWeight' : 'float64',
            'DeltaXCcyBasisRiskWeight' : 'float64',
            'DeltaTenorRhoTheta' : 'float64',
            'VegaUnderlyingRhoAlpha' : 'float64',
            'DeltaTenorRho' : 'float64',
            'DeltaCurveRho' : 'float64',
            'DeltaInflationRho' : 'float64',
            'DeltaXCcyBasisRho' : 'float64',
            'Gamma' : 'float64'
        },
        'MS_CR' : {
            'Bucket' : 'str',
            'CoveredBondBucket' :  'str',
            'OtherBucket' : 'str',
            'IndexBuckets' : 'str',
            'DeltaBucketRiskWeight' : 'float64',
            'DeltaCovBondAARiskWeight' : 'float64',
            'DeltaNameRho' : 'float64',
            'DeltaTenorRho' : 'float64',
            'DeltaBasisRho' : 'float64',
            'DeltaNameIndexRho' : 'float64',
            'DeltaTenorIndexRho' : 'float64',
            'DeltaBasisIndexRho' : 'float64',
            'Gamma' : 'float64',
        },
        'MS_CC' : {
            'Bucket' : 'str',
            'OtherBucket' : 'str',
            'DeltaBucketRiskWeight' : 'float64',
            'DeltaNameRho' : 'float64',
            'DeltaTenorRho' : 'float64',
            'DeltaBasisRho' : 'float64',
            'Gamma' : 'float64'
        },
        'MS_CS' : {
            'Bucket' : 'str',
            'OtherBucket' : 'str',
            'DeltaBucketRiskWeight' : 'float64',
            'DeltaTrancheRho' : 'float64',
            'DeltaTenorRho' : 'float64',
            'DeltaBasisRho' : 'float64',
            'Gamma' : 'float64'
        },
        'MS_EQ' : {
            'Bucket' : 'str',
            'OtherBucket' : 'str',
            'MarketCapThreshold' : 'int64',
            'DeltaBucketRiskWeight' : 'float64',
            'DeltaNameBucketRho' : 'float64',
            'DeltaSpotRepoRho' : 'float64',
            'Gamma' : 'float64'
        },
        'MS_CM' : {
            'Bucket' : 'str',
            'DeltaBucketRiskWeight' : 'float64',
            'DeltaCommodityRho' : 'float64',
            'DeltaTenorRho' : 'float64',
            'DeltaBasisRho' : 'float64',
            'Gamma' : 'float64',
        },
        'MS_FX' : {
            'DeltaRiskWeight' : 'float64',
            'Gamma' : 'float64',
            'ERMIIBand' : 'float64',
            'ERMIICcys' : 'float64',
        },
        'MD_CR' : {
            'LGDSenior' : 'float64',
            'LGDCovered' : 'float64',
            'CQRiskWeight' : 'float64'
        },
        'MD_CC' : {
            'CQRiskWeight' : 'float64'
        },
        'MD_CS' : {
            'CQRiskWeight' : 'float64'
        },
        'MR_RR' : {
            'RiskWeight' : 'float64'
        },
        'CVA' : {
            'BA-Bucket' : 'str',
            'BA-Rho' : 'float64',
            'BA-DiscountFactor' : 'float64',
            'BA-Alpha' : 'float64',
            'BA-Beta' : 'float64',
            'BA-DiscountScalar' : 'float64',
            'BA-RiskWeight' : 'float64',
            'BA-rDirect' : 'float64',
            'BA-rRelated' : 'float64',
            'BA-rSectorRegion' : 'float64',
            'BA-IndexDiversifictaion' : 'float64',
            'SA-HedgeDisallowance' : 'float64',
            'SA-CapitalMultiplier' : 'float64',
        },
        'CS_IR' : {
            'DeltaTenorRiskWeight' : 'float64',
            'DeltaInflationRiskWeight' : 'float64',
            'DeltaTenorIlliquidRiskWeight' : 'float64',
            'DeltaInflationIlliquidRiskWeight' : 'float64',
            'DeltaTenorRho' : 'float64',
            'DeltaInflationRho' : 'float64',
            'DeltaIlliquidRho' : 'float64',
            'VegaRiskWeight' : 'float64',
            'VegaRho' : 'float64',
            'Gamma' : 'float64',
        },
        'CS_FX' : {
            'DeltaRiskWeight' : 'float64',
            'VegaRiskWeight' : 'float64',
            'Gamma' : 'float64',
            'ERMIIBand' : 'float64',
            'ERMIICcys' : 'float64',
        },
        'CS_CC' : {
            'Bucket' : 'str',
            'CoveredBondBucket' :  'str',
            'IndexBuckets' : 'str',
            'DeltaRiskWeight' : 'float64',
            'DeltaNameRelatedRho' : 'float64',
            'DeltaNameUnrelatedRho' : 'float64',
            'DeltaTenorRho' : 'float64',
            'DeltaCreditQualityRho' : 'float64',
            'DeltaNameRelatedIndexRho' : 'float64',
            'DeltaNameUnrelatedIndexRho' : 'float64',
            'DeltaTenorIndexRho' : 'float64',
            'DeltaCreditQualityIndexRho' : 'float64',
            'Gamma' : 'float64'
        },
        'CS_CR' : {
            'Bucket' : 'str',
            'DeltaBucketRiskWeight' : 'float64',
            'VegaRiskWeight' : 'float64',
            'Gamma' : 'float64'
        },
        'CS_EQ' : {
            'Bucket' : 'str',
            'OtherBucket' : 'str',
            'MarketCapThreshold' : 'int64',
            'DeltaBucketRiskWeight' : 'float64',
            'VegaBucketRiskWeight' : 'float64',
            'DeltaBucketRho' : 'float64',
            'Gamma' : 'float64'
        },
        'CS_CM' : {
            'Bucket' : 'str',
            'DeltaBucketRiskWeight' : 'float64',
            'VegaRiskWeight' : 'float64',
            'Gamma' : 'float64'
        }
    }


    def __init__(self, regulator):
        self._name = type(self).__name__
        self._regulator = regulator
        self._config = self.readConfig()
        self._computeVegaRiskWeights()
        self._computeRho()


    def _computeVegaRiskWeights(self):
        if 'MR' not in self._config.keys():
            return

        cfg = self._config['MR']

        if 'VegaLiquidityHorizon' not in cfg.keys():
            return

        vegaLH = cfg['VegaLiquidityHorizon']
        RWSigma = cfg['VegaRiskWeightSigma']
        sqrtTen = 10 ** 0.5
        vegaLH.loc[:, 'RiskWeight'] = vegaLH['LiquidityHorizon'].astype('float64').apply(lambda x: min(RWSigma * (x ** 0.5) / sqrtTen, 1))
        self._config['MR']['VegaLiquidityHorizon'] = vegaLH

        for key, grp in vegaLH.groupby('AssetClass'):
            if key in self._config.keys():
                if key == 'MS_EQ':
                    grp.drop(columns=['AssetClass', 'LiquidityHorizon'], inplace=True)
                    grp.set_index('MarketCap', inplace=True)
                    self._config[key]['VegaRiskWeight'] = grp
                else:
                    self._config[key]['VegaRiskWeight'] = grp['RiskWeight'].iat[0]

    def __rhoExpr(self, tau, a, b):
        return math.exp(-tau * abs(a - b) / min(a, b))

    def _computeRho(self):
        cfg = self._config['MS_IR']

        if 'DeltaTenorRhoTheta' in cfg.keys() and 'DeltaTenors' in cfg.keys():
            theta = cfg['DeltaTenorRhoTheta']
            rho = np.ones((len(cfg['DeltaTenors']), len(cfg['DeltaTenors'])))

            for i, r in enumerate(cfg['DeltaTenors']):
                for j, c in enumerate(cfg['DeltaTenors']):
                    if j <= i:
                        continue
                    else:
                        rho[i, j] = rho [j, i] = max(self.__rhoExpr(theta, float(r), float(c)), 0.4)

            cfg['DeltaTenorRho'] = pd.DataFrame(rho, index=cfg['DeltaTenors'], columns=cfg['DeltaTenors'])

        if 'VegaTenors' in cfg.keys():
            alpha = cfg['VegaUnderlyingRhoAlpha']
            rho = np.ones((len(cfg['VegaTenors']), len(cfg['VegaTenors'])))

            for i, r in enumerate(cfg['VegaTenors']):
                for j, c in enumerate(cfg['VegaTenors']):
                    if j <= i:
                        continue
                    else:
                        rho[i, j] = rho [j, i] = self.__rhoExpr(alpha, float(r), float(c))

            cfg['VegaUnderlyingTenorRho'] = pd.DataFrame(rho, index=cfg['VegaTenors'], columns=cfg['VegaTenors'])

        self._config['MS_IR'] = cfg
        alpha = self._config['MR']['VegaOptionRhoAlpha']

        for assetClass in ['MS_IR', 'MS_CR', 'MS_CC', 'MS_CS', 'MS_EQ', 'MS_CM', 'MS_FX']:
            cfg = self._config[assetClass]

            if 'VegaTenors' in cfg.keys():
                rho = np.ones((len(cfg['VegaTenors']), len(cfg['VegaTenors'])))

                for i, r in enumerate(cfg['VegaTenors']):
                    for j, c in enumerate(cfg['VegaTenors']):
                        if j <= i:
                            continue
                        else:
                            rho[i, j] = rho [j, i] = self.__rhoExpr(alpha, float(r), float(c))

                cfg['VegaOptionTenorRho'] = pd.DataFrame(rho, index=cfg['VegaTenors'], columns=cfg['VegaTenors'])
                self._config[assetClass] = cfg


    def getCellValues(self, ws):
        for r in ws.iter_rows():
            vals = []
            nulls = 0

            for c in r:
                if c.value is None:
                    nulls += 1
                else:
                    if nulls:
                        vals.extend([''] * nulls)
                        nulls = 0

                    vals.append(str(c.value))

            yield vals


    def readConfig(self):
        """
        Read configuration from file, preferring JSON format over Excel.

        Searches for config files in the Configs directory, trying JSON first,
        then falling back to Excel if JSON is not found.
        """
        cfpath = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'FRTB', 'Configs')

        # Try JSON first (preferred format)
        jsonfile = os.path.join(cfpath, self._configFileJSON.format(self._regulator))
        if os.path.exists(jsonfile):
            return self._readConfigFromJSON(jsonfile)

        # Fall back to Excel
        xlfile = os.path.join(cfpath, self._configFileExcel.format(self._regulator))
        if os.path.exists(xlfile):
            return self._readConfigFromExcel(xlfile)

        raise FileNotFoundError(f"No config file found for regulator '{self._regulator}' in {cfpath}")

    def _readConfigFromExcel(self, xlfile: str) -> Dict[str, Any]:
        """Read configuration from an Excel file."""
        wb = xl.load_workbook(xlfile, read_only=True, data_only=True)
        cfg = {}

        for i, ws in enumerate(wb.worksheets):
            configClass = wb.sheetnames[i]

            if configClass in self._riskClassKeyDataType.keys():
                cfgdf = pd.DataFrame(self.getCellValues(ws)).fillna('')
                cfg[configClass] = FNU.extractKeyedData(configClass, cfgdf, self._riskClassKeyDataType[configClass], **self._riskClassCongigKeyTypes[configClass])
            elif configClass != 'Copyright':
                print(f"Unknown config sheet : {configClass} in config for {self._regulator} in {xlfile}")

        wb.close()
        return cfg

    def _readConfigFromJSON(self, jsonfile: str) -> Dict[str, Any]:
        """
        Read configuration from a JSON file.

        Converts the self-documenting JSON format back to the internal
        representation using pandas DataFrames and Series.
        """
        with open(jsonfile, 'r') as f:
            raw_config = json.load(f)

        cfg = {}

        for risk_class, data in raw_config.items():
            # Skip copyright and other metadata
            if risk_class.startswith('_'):
                continue

            if risk_class not in self._riskClassKeyDataType.keys():
                print(f"Unknown config section: {risk_class} in {jsonfile}")
                continue

            cfg[risk_class] = self._parseJSONRiskClass(risk_class, data)

        return cfg

    def _parseJSONRiskClass(self, risk_class: str, data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Parse a risk class section from JSON into the internal format.

        Converts JSON structures back to pandas DataFrames/Series with proper
        indices and data types.
        """
        result = {}
        key_types = self._riskClassCongigKeyTypes.get(risk_class, {})
        data_types = self._riskClassKeyDataType.get(risk_class, {})

        add_index = key_types.get('addIndex', {})
        add_columns = key_types.get('addColumns', {})

        for key, item in data.items():
            if not isinstance(item, dict) or 'type' not in item:
                # Skip malformed entries
                continue

            item_type = item['type']
            value = item.get('value')
            dtype = item.get('dtype', 'str')

            if item_type == 'scalar':
                # Convert scalar to appropriate type
                if dtype in self._TYPE_MAP:
                    result[key] = self._TYPE_MAP[dtype](value)
                else:
                    result[key] = value

            elif item_type == 'list':
                # Convert to pandas Series
                values = value
                if dtype in self._TYPE_MAP and dtype != 'str':
                    values = [self._TYPE_MAP[dtype](v) if v != '' else self._FILLNA_MAP[dtype] for v in value]

                index = item.get('index', None)
                result[key] = pd.Series(values, index=index, name=key)

                # Apply dtype
                if dtype in data_types.get(key, dtype):
                    try:
                        result[key] = result[key].astype(dtype)
                    except (ValueError, TypeError):
                        pass

            elif item_type == 'dataframe':
                # Convert to pandas DataFrame
                columns = item.get('columns', [])
                dtypes = item.get('dtypes', {})
                index = item.get('index', None)

                df = pd.DataFrame(value)

                # Ensure column order matches
                if columns:
                    df = df[columns]

                # Apply dtypes
                for col, col_dtype in dtypes.items():
                    if col in df.columns and col_dtype in self._TYPE_MAP:
                        try:
                            df[col] = df[col].astype(col_dtype)
                        except (ValueError, TypeError):
                            pass

                # Set index if provided
                if index is not None:
                    df.index = index

                result[key] = df

        # Apply addIndex expressions only if index wasn't already set from JSON
        for key, expr in add_index.items():
            if key in result:
                # Skip if index was already set from JSON (not a RangeIndex)
                if not isinstance(result[key].index, pd.RangeIndex):
                    continue
                try:
                    result[key].index = eval(expr, {'dataDict': result})
                except Exception as err:
                    print(f"Warning: Error setting index on '{key}': {err}")

        # Apply addColumns expressions only if columns weren't already set from JSON
        for key, expr in add_columns.items():
            if key in result and isinstance(result[key], pd.DataFrame):
                # Skip - columns are already properly named from JSON
                pass

        return result

    def writeConfigToJSON(self, json_path: str = None, pretty: bool = True, indent: int = 2) -> str:
        """
        Write the current configuration to a JSON file.

        Args:
            json_path: Path to save JSON file (optional, defaults to Configs directory)
            pretty: Whether to pretty-print the JSON
            indent: Indentation level for pretty printing

        Returns:
            Path to the created JSON file
        """
        if json_path is None:
            cfpath = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'FRTB', 'Configs')
            json_path = os.path.join(cfpath, self._configFileJSON.format(self._regulator))

        json_config = self._configToJSONFormat()

        with open(json_path, 'w') as f:
            if pretty:
                json.dump(json_config, f, indent=indent, cls=NumpyEncoder)
            else:
                json.dump(json_config, f, cls=NumpyEncoder)

        return json_path

    def _configToJSONFormat(self) -> Dict[str, Any]:
        """
        Convert the internal configuration to the self-documenting JSON format.

        Returns:
            Dictionary suitable for JSON serialization
        """
        json_config = {
            '_copyright': {
                'value': [
                    f'Copyright (C) 2024-2025 frtb.net limited',
                    '',
                    'Contact us at <info@frtb.net> or via our website at <https://frtb.net>',
                    '',
                    '',
                    'This program is free software: you can redistribute it and/or modify',
                    'it under the terms of the GNU Affero General Public License as',
                    'published by the Free Software Foundation, either version 3 of the',
                    'License, or (at your option) any later version.',
                    '',
                    'This program is distributed in the hope that it will be useful,',
                    'but WITHOUT ANY WARRANTY; without even the implied warranty of',
                    'MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the',
                    'GNU Affero General Public License for more details.',
                    '',
                    'You should have received a copy of the GNU Affero General Public License',
                    'along with this program. If not, see <https://www.gnu.org/licenses/>.'
                ],
                'type': 'text',
                'note': 'Copyright and license information'
            }
        }

        for risk_class, data in self._config.items():
            json_config[risk_class] = self._riskClassToJSONFormat(risk_class, data)

        return json_config

    def _riskClassToJSONFormat(self, risk_class: str, data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Convert a risk class data dictionary to JSON format.

        Args:
            risk_class: Name of the risk class
            data: Dictionary of config items for this risk class

        Returns:
            Dictionary with self-documenting JSON structure
        """
        result = {}
        data_types = self._riskClassKeyDataType.get(risk_class, {})

        for key, value in data.items():
            result[key] = self._valueToJSONFormat(key, value, data_types.get(key, None))

        return result

    def _valueToJSONFormat(self, key: str, value: Any, dtype_hint: str = None) -> Dict[str, Any]:  # noqa: ARG002
        """
        Convert a single value to the self-documenting JSON format.

        Args:
            key: Name of the config item
            value: The value to convert
            dtype_hint: Optional type hint from the data types dict

        Returns:
            Dictionary with type, value, and dtype fields
        """
        if isinstance(value, pd.DataFrame):
            # DataFrame
            # Convert column names to strings for JSON compatibility
            str_columns = [str(c) for c in value.columns]
            dtypes = {}
            for col in value.columns:
                col_dtype = str(value[col].dtype)
                str_col = str(col)
                if col_dtype == 'object':
                    dtypes[str_col] = 'str'
                elif 'int' in col_dtype:
                    dtypes[str_col] = 'int64'
                elif 'float' in col_dtype:
                    dtypes[str_col] = 'float64'
                else:
                    dtypes[str_col] = col_dtype

            # Convert DataFrame with string column names for JSON
            value_renamed = value.copy()
            value_renamed.columns = str_columns

            result = {
                'value': value_renamed.to_dict(orient='list'),
                'type': 'dataframe',
                'columns': str_columns,
                'dtypes': dtypes
            }

            # Include index if it's not the default RangeIndex
            if not isinstance(value.index, pd.RangeIndex):
                result['index'] = [str(i) for i in value.index.tolist()]
                if value.index.name:
                    result['index_name'] = str(value.index.name)

            return result

        elif isinstance(value, pd.Series):
            # Series (list)
            dtype = str(value.dtype)
            if dtype == 'object':
                dtype = 'str'
            elif 'int' in dtype:
                dtype = 'int64'
            elif 'float' in dtype:
                dtype = 'float64'

            result = {
                'value': value.tolist(),
                'type': 'list',
                'dtype': dtype,
                'name': key
            }

            # Include index if it's not the default RangeIndex
            if not isinstance(value.index, pd.RangeIndex):
                result['index'] = [str(i) for i in value.index.tolist()]

            return result

        else:
            # Scalar value
            if isinstance(value, (np.integer, int)):
                dtype = 'int64'
                value = int(value)
            elif isinstance(value, (np.floating, float)):
                dtype = 'float64'
                value = float(value)
            elif isinstance(value, bool):
                dtype = 'bool'
            else:
                dtype = 'str'
                value = str(value)

            return {
                'value': value,
                'type': 'scalar',
                'dtype': dtype
            }

    @classmethod
    def excelToJSON(cls, excel_path: str, json_path: str = None,
                    pretty: bool = True, indent: int = 2) -> Dict[str, Any]:
        """
        Convert an Excel config file to JSON format (class method).

        This is a standalone conversion that doesn't require loading the config
        into the FRTBConfig object first.

        Args:
            excel_path: Path to the Excel file
            json_path: Path to save JSON file (optional, defaults to same name with .json)
            pretty: Whether to pretty-print the JSON
            indent: Indentation level for pretty printing

        Returns:
            Dictionary containing the converted configuration
        """
        excel_path = Path(excel_path)
        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        if json_path is None:
            json_path = excel_path.with_suffix('.json')
        else:
            json_path = Path(json_path)

        # Extract regulator from filename
        stem = excel_path.stem
        if '_' in stem:
            regulator = stem.split('_', 1)[1]
        else:
            regulator = 'UNKNOWN'

        print(f"Converting Excel → JSON: {excel_path} → {json_path}")

        # Create a temporary instance to use its parsing logic
        # We temporarily bypass normal init to avoid loading the config
        instance = object.__new__(cls)
        instance._name = cls.__name__
        instance._regulator = regulator

        # Read from Excel
        instance._config = instance._readConfigFromExcel(str(excel_path))

        # Compute derived values
        instance._computeVegaRiskWeights()
        instance._computeRho()

        # Write to JSON
        json_config = instance._configToJSONFormat()

        with open(json_path, 'w') as f:
            if pretty:
                json.dump(json_config, f, indent=indent, cls=NumpyEncoder)
            else:
                json.dump(json_config, f, cls=NumpyEncoder)

        print(f"Conversion complete: {json_path}")
        return json_config

    @classmethod
    def jsonToExcel(cls, json_path: str, excel_path: str = None) -> str:
        """
        Convert a JSON config file to Excel format (class method).

        Args:
            json_path: Path to the JSON file
            excel_path: Path to save Excel file (optional)

        Returns:
            Path to the created Excel file
        """
        json_path = Path(json_path)
        if not json_path.exists():
            raise FileNotFoundError(f"JSON file not found: {json_path}")

        if excel_path is None:
            excel_path = json_path.with_suffix('.xlsx')
        else:
            excel_path = Path(excel_path)

        print(f"Converting JSON → Excel: {json_path} → {excel_path}")

        # Load JSON
        with open(json_path, 'r') as f:
            config = json.load(f)

        # Create Excel workbook
        wb = xl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Add copyright sheet
        if '_copyright' in config:
            ws_copyright = wb.create_sheet('Copyright')
            copyright_data = config['_copyright']
            if copyright_data.get('type') == 'text' and 'value' in copyright_data:
                for row_num, line in enumerate(copyright_data['value'], start=1):
                    ws_copyright.cell(row_num, 1, line)
        else:
            ws_copyright = wb.create_sheet('Copyright')
            ws_copyright['A1'] = '© 2024 frtb.net limited'
            ws_copyright['A2'] = 'Licensed under GNU Affero General Public License v3.0'

        # Process each risk class
        for risk_class, data in config.items():
            if risk_class == '_copyright':
                continue
            print(f"  Creating sheet: {risk_class}")
            ws = wb.create_sheet(risk_class)
            cls._writeSheetData(ws, data)

        wb.save(excel_path)
        print(f"Conversion complete: {excel_path}")

        return str(excel_path)

    @staticmethod
    def _writeSheetData(ws, data: Dict[str, Any]):
        """Write data dictionary to an Excel worksheet."""
        row_num = 1

        for key, item in data.items():
            if not isinstance(item, dict) or 'type' not in item:
                continue

            item_type = item['type']

            if item_type == 'scalar':
                ws.cell(row_num, 1, key)
                ws.cell(row_num, 2, item['value'])
                row_num += 1

            elif item_type == 'list':
                ws.cell(row_num, 1, key)
                for col_idx, val in enumerate(item['value'], start=2):
                    ws.cell(row_num, col_idx, val)
                row_num += 1

            elif item_type == 'dataframe':
                columns = item['columns']
                has_index = 'index' in item and item['index']

                # Write header row
                ws.cell(row_num, 1, key)
                col_start = 2

                if has_index:
                    index_name = item.get('index_name', '')
                    ws.cell(row_num, 2, index_name)
                    col_start = 3

                for col_idx, col_name in enumerate(columns, start=col_start):
                    ws.cell(row_num, col_idx, col_name)
                row_num += 1

                # Write data rows
                df_dict = item['value']
                num_rows = len(df_dict[columns[0]]) if columns else 0

                for row_idx in range(num_rows):
                    col_num = 2

                    if has_index:
                        ws.cell(row_num, 2, item['index'][row_idx])
                        col_num = 3

                    for col_name in columns:
                        val = df_dict[col_name][row_idx]
                        ws.cell(row_num, col_num, val)
                        col_num += 1

                    row_num += 1

    def _argCheck(self, riskClass, item=None):
        if not riskClass in self._config.keys():
            raise ValueError(f"{self._name}: no config for riskClass '{riskClass}'")

        if not item is None and not item in self._config[riskClass].keys():
            raise ValueError(f"{self._name}: no config item '{item}' for riskClass '{riskClass}'")


    def getConfigList(self):
        return self._config.keys()

    def getConfig(self, riskClass):
        self._argCheck(riskClass)
        return self._config[riskClass]

    def getConfigItem(self, riskClass, item):
        self._argCheck(riskClass, item)
        return self._config[riskClass][item]

    def getBuckets(self, riskClass, buckets=None):
        if riskClass == 'CVA':
            self._argCheck(riskClass, 'BA-Bucket')
            bdf = self._config[riskClass]['BA-Bucket']
        else:
            self._argCheck(riskClass, 'Bucket')
            bdf = self._config[riskClass]['Bucket']

        if buckets is None:
            if isinstance(bdf, pd.DataFrame):
                return bdf['Bucket'].unique().tolist()
            else:
                return bdf.unique().tolist()
        else:
            if isinstance(bdf, pd.DataFrame()):
                return bdf[bdf['Bucket'].isin(buckets)]['Bucket'].to_list()
            else:
                return bdf[bdf['Bucket'].isin(buckets)].to_list()


if __name__ == '__main__':
    config = FRTBConfig('BCBS')
    # print(config.getConfigItem('MS_CS', 'DeltaBucketRiskWeight'))

    for cfg in config.getConfigList():
        print
        print('=' * len(cfg))
        print(cfg)
        print('=' * len(cfg))
        print(config.getConfig(cfg))

        if not cfg in ['MR', 'MS_IR', 'MS_FX', 'MD_CC', 'CS_IR', 'CS_FX']:
            print(config.getBuckets(cfg))

    print(config.getConfig('MS_FX'))
